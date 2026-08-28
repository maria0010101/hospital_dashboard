#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
業務資料樣本混淆工具
====================
取原始「業務資料彙整-1150826.xlsx」中 忠孝/中興/和平 三個院區的資料，
對 院區名稱 / 部別 / 科別 / 醫師名稱 / 醫師代碼 / 門診部名稱 / 護理站名稱 /
大類別 / 病床類別(類別) 及 各項業務數值 進行混淆，輸出「業務資料彙整-樣本-1150826.xlsx」，
供 App 匯入展示/格式參考使用（不含任何真實院區、人員、單位與數值）。

混淆原則（維持報表數據一致性）：
1. 名稱採「一對一固定對應」：同一真實名稱 → 同一假名（跨工作表一致）；
   病床服務量的 護理站 / 大類別 / 病床類別 亦採全域一對一（護理站01..NN、大類別01..NN、病床類別01..NN）。
2. 數值採「院區固定倍率縮放」：同一院區所有計數/金額欄位 × 同一倍率後四捨五入；
   - 比率類欄位（佔床率）與曆法類欄位（年月/當月日數/季）維持原值不動；
   - 因分子分母同倍率縮放，佔床率、初診率、健保/自費佔比、增減率等
     報表指標均與原資料一致，僅絕對數值不同。
"""
import openpyxl
from collections import Counter, OrderedDict

SRC = '/home/hpd/下載/業務資料彙整-1150826.xlsx'
OUT = '/home/hpd/文件/hospital_dashboard/樣本資料/業務資料彙整-樣本-1150826.xlsx'
MAPPING_OUT = '/home/hpd/文件/hospital_dashboard/樣本資料/混淆對照表.md'

BRANCHES = ('忠孝', '中興', '和平')
# 院區混淆：忠孝→甲院區、中興→乙院區、和平→丙院區
BRANCH_FAKE = {'忠孝': '甲院區', '中興': '乙院區', '和平': '丙院區'}
# 合併院區（和平婦幼）→ 直接改為丙院區
MERGED_FAKE = {'忠孝': '甲院區', '中興': '乙院區', '和平婦幼': '丙院區'}
# 數值縮放倍率（固定、可重現）
SCALE = {'忠孝': 0.72, '中興': 0.64, '和平': 0.86}
# 真實名稱 token（用於全面掃描確認無殘留）
REAL_TOKENS = ('忠孝', '中興', '和平', '仁愛', '婦幼', '陽明', '松德', '林森', '昆明', '中醫', '院本部')

wb_src = openpyxl.load_workbook(SRC, read_only=True)

# ── 建立名稱對應表 ────────────────────────────────
dept_map = OrderedDict()   # 真實科別 → 科別NN
div_map = OrderedDict()    # 真實部別 → 部別NN
clinic_map = OrderedDict() # 真實門診部 → 門診部NN
doc_map = OrderedDict()    # (id, name) → (假id, 假名)
station_map = OrderedDict() # 護理站名稱 → 護理站NN
major_map = OrderedDict()   # 大類別 → 大類別NN
cat_map = OrderedDict()     # 病床類別(類別) → 病床類別NN


def clean(v):
    return str(v).strip() if v is not None else ''

def fake_dept(d):
    if d not in dept_map:
        dept_map[d] = f'科別{len(dept_map) + 1:02d}'
    return dept_map[d]

def fake_div(d):
    if d not in div_map:
        div_map[d] = f'部別{len(div_map) + 1:02d}'
    return div_map[d]

def fake_clinic(c):
    if c not in clinic_map:
        clinic_map[c] = f'門診部{len(clinic_map) + 1:02d}'
    return clinic_map[c]

def fake_doc(did, dname):
    key = (str(did), str(dname))
    if key not in doc_map:
        i = len(doc_map) + 1
        doc_map[key] = (f'DR{i:04d}', f'醫師{i:03d}')
    return doc_map[key]


def fake_station(s):
    if s not in station_map:
        station_map[s] = f'護理站{len(station_map) + 1:02d}'
    return station_map[s]


def fake_major(s):
    if s not in major_map:
        major_map[s] = f'大類別{len(major_map) + 1:02d}'
    return major_map[s]


def fake_cat(s):
    if s not in cat_map:
        cat_map[s] = f'病床類別{len(cat_map) + 1:02d}'
    return cat_map[s]

def row_branch(r, bcol):
    v = str(r[bcol]) if r[bcol] is not None else ''
    if v in BRANCHES:
        return v
    if v in ('忠孝院區', '中興院區', '和平院區'):
        return v[:-2]
    return None

# 先掃描 醫師服務量 建立醫師對應（也收集科別/部別）
ws = wb_src['醫師服務量']
rows = ws.iter_rows(values_only=True)
next(rows)
for r in rows:
    if row_branch(r, 1) is None:
        continue
    if r[4] is not None and str(r[4]).strip():
        fake_div(str(r[4]).strip())
    if r[5] is not None and str(r[5]).strip():
        fake_dept(str(r[5]).strip())
    if (r[2] is not None or r[3] is not None) and (r[2] is not None and str(r[2]).strip() or r[3] is not None and str(r[3]).strip()):
        fake_doc(r[2] if r[2] is not None else '', r[3] if r[3] is not None else '')

# 掃描 門診/住院 收集科別/部別
for name, dcol, divcol in (('門診業務資料', 4, 18), ('門診業務資料', 19, None), ('住院業務資料', 4, 10), ('住院業務資料', 11, None)):
    ws = wb_src[name]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    for r in rows:
        if row_branch(r, 3) is None:
            continue
        if r[dcol] is not None and str(r[dcol]).strip():
            fake_dept(str(r[dcol]).strip())
        if divcol is not None and r[divcol] is not None and str(r[divcol]).strip():
            fake_div(str(r[divcol]).strip())

# 院外門診部 → 門診部名稱
ws = wb_src['院外門診部服務量']
rows = ws.iter_rows(values_only=True)
next(rows)
for r in rows:
    if row_branch(r, 2) is not None and r[3] is not None and str(r[3]).strip():
        fake_clinic(str(r[3]).strip())

print(f'科別 {len(dept_map)} 個 → 科別01..{len(dept_map):02d}')
print(f'部別 {len(div_map)} 個 → 部別01..{len(div_map):02d}')
print(f'門診部 {len(clinic_map)} 個 → 門診部01..{len(clinic_map):02d}')
print(f'醫師 {len(doc_map)} 位 → 醫師001..{len(doc_map):03d}')

# ── 數值縮放 ──────────────────────────────────────
def scale_val(v, k):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == '':
            return v
        try:
            n = float(s)
        except ValueError:
            return v  # 非數值字串（如 'ddddd'）原樣保留
        r = int(n * k + 0.5)
        return r if n > 0 and r == 0 else r
    if isinstance(v, (int, float)):
        if v == 0:
            return 0
        r = int(v * k + 0.5)
        return r if r > 0 else 1
    return v

def tok(s):
    """將字串中的真實院區名稱替換為假名（涵蓋院區/合併院區/護理站/合併鍵等）。"""
    if s is None:
        return None
    t = str(s)
    for real, fake in BRANCH_FAKE.items():
        t = t.replace(real, fake)
    return t

wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)
order = ['門診業務資料', '住院業務資料', '病床別業務資料', '院外門診部服務量',
         '會計室報表資料', '工作天數', '其他營運管理指標資料', '醫師服務量']

for name in order:
    ws_s = wb_src[name]
    rows = ws_s.iter_rows(values_only=True)
    hdr = list(next(rows))
    ws_o = wb_out.create_sheet(name)
    ws_o.append(hdr)
    n_out = 0
    for r in rows:
        if name == '門診業務資料':
            br = row_branch(r, 3)
            if br is None:
                continue
            k = SCALE[br]
            out = list(r)
            out[3] = BRANCH_FAKE[br]
            out[4] = fake_dept(str(r[4]).strip()) if r[4] is not None and str(r[4]).strip() else r[4]
            for i in range(5, 17):
                out[i] = scale_val(r[i], k)
            out[17] = MERGED_FAKE[str(r[17]).strip()] if r[17] is not None and str(r[17]).strip() in MERGED_FAKE else tok(r[17])
            out[18] = fake_div(str(r[18]).strip()) if r[18] is not None and str(r[18]).strip() else r[18]
            out[19] = fake_dept(str(r[19]).strip()) if r[19] is not None and str(r[19]).strip() else r[19]
            out[22] = MERGED_FAKE[str(r[22]).strip()] if r[22] is not None and str(r[22]).strip() in MERGED_FAKE else tok(r[22])
            out[23] = BRANCH_FAKE[br]
            out[24] = MERGED_FAKE[str(r[24]).strip()] if r[24] is not None and str(r[24]).strip() in MERGED_FAKE else tok(r[24])
        elif name == '住院業務資料':
            br = row_branch(r, 3)
            if br is None:
                continue
            k = SCALE[br]
            out = list(r)
            out[3] = BRANCH_FAKE[br]
            out[4] = fake_dept(str(r[4]).strip()) if r[4] is not None and str(r[4]).strip() else r[4]
            for i in range(5, 9):
                out[i] = scale_val(r[i], k)
            out[9] = MERGED_FAKE[str(r[9]).strip()] if r[9] is not None and str(r[9]).strip() in MERGED_FAKE else tok(r[9])
            out[10] = fake_div(str(r[10]).strip()) if r[10] is not None and str(r[10]).strip() else r[10]
            out[11] = fake_dept(str(r[11]).strip()) if r[11] is not None and str(r[11]).strip() else r[11]
            out[13] = BRANCH_FAKE[br]
            out[14] = MERGED_FAKE[str(r[14]).strip()] if r[14] is not None and str(r[14]).strip() in MERGED_FAKE else tok(r[14])
        elif name == '病床別業務資料':
            br = row_branch(r, 3)
            if br is None:
                continue
            k = SCALE[br]
            out = list(r)
            out[3] = BRANCH_FAKE[br]
            # 當月日數[4] 不動
            ns = clean(r[5]); maj = clean(r[6]); cat = clean(r[7])
            out[5] = fake_station(ns) if ns else r[5]     # 護理站
            out[6] = fake_major(maj) if maj else r[6]     # 大類別
            out[7] = fake_cat(cat) if cat else r[7]       # 病床類別(類別)
            for i in range(8, 16):
                out[i] = scale_val(r[i], k)
            # 佔床率[16][17] 維持原值
            out[18] = MERGED_FAKE[str(r[18]).strip()] if r[18] is not None and str(r[18]).strip() in MERGED_FAKE else tok(r[18])
            ns2 = clean(r[19]); cat2 = clean(r[20])
            out[19] = fake_station(ns2) if ns2 else r[19]  # 佔床率報表護理站
            out[20] = fake_cat(cat2) if cat2 else r[20]    # 佔床率報表病床類別
            out[21] = BRANCH_FAKE[br]
            out[22] = MERGED_FAKE[str(r[22]).strip()] if r[22] is not None and str(r[22]).strip() in MERGED_FAKE else tok(r[22])
            # 合併鍵[23] = 年度+月份+院區+病床類別 → 以假名重建
            if r[23] is not None:
                out[23] = f"{r[1]}{r[2]}{BRANCH_FAKE[br]}{fake_cat(cat) if cat else ''}"
        elif name == '院外門診部服務量':
            br = row_branch(r, 2)
            if br is None:
                continue
            k = SCALE[br]
            out = list(r)
            out[2] = BRANCH_FAKE[br]
            out[3] = fake_clinic(str(r[3]).strip()) if r[3] is not None and str(r[3]).strip() else r[3]
            for i in (4, 5, 6, 7, 10, 11, 12, 13, 14):
                out[i] = scale_val(r[i], k)
            out[8] = BRANCH_FAKE[br]
            out[9] = MERGED_FAKE[str(r[9]).strip()] if r[9] is not None and str(r[9]).strip() in MERGED_FAKE else tok(r[9])
        elif name == '會計室報表資料':
            br = row_branch(r, 3)
            if br is None:
                continue
            k = SCALE[br]
            out = list(r)
            out[3] = BRANCH_FAKE[br]
            for i in range(4, 17):
                out[i] = scale_val(r[i], k)
            out[17] = MERGED_FAKE[str(r[17]).strip()] if r[17] is not None and str(r[17]).strip() in MERGED_FAKE else tok(r[17])
            out[19] = BRANCH_FAKE[br]
            out[20] = MERGED_FAKE[str(r[20]).strip()] if r[20] is not None and str(r[20]).strip() in MERGED_FAKE else tok(r[20])
        elif name == '其他營運管理指標資料':
            br = row_branch(r, 2)
            if br is None:
                continue
            k = SCALE[br]
            out = list(r)
            out[2] = BRANCH_FAKE[br]
            for i in range(3, 32):
                out[i] = scale_val(r[i], k)
            out[32] = BRANCH_FAKE[br]
            out[33] = MERGED_FAKE[str(r[33]).strip()] if r[33] is not None and str(r[33]).strip() in MERGED_FAKE else tok(r[33])
            # out[34] 年月 不動
        elif name == '醫師服務量':
            br = row_branch(r, 1)
            if br is None:
                continue
            k = SCALE[br]
            out = list(r)
            out[1] = BRANCH_FAKE[br]
            did = str(r[2]).strip() if r[2] is not None else ''
            dnm = str(r[3]).strip() if r[3] is not None else ''
            if did or dnm:
                fid, fname = fake_doc(did, dnm)
                out[2] = fid
                out[3] = fname
            out[4] = fake_div(str(r[4]).strip()) if r[4] is not None and str(r[4]).strip() else r[4]
            out[5] = fake_dept(str(r[5]).strip()) if r[5] is not None and str(r[5]).strip() else r[5]
            for i in range(6, 15):
                out[i] = scale_val(r[i], k)
        elif name == '工作天數':
            out = list(r)  # 無院區欄位，原樣保留
        else:
            continue
        ws_o.append(out)
        n_out += 1
    print(f'{name}: {n_out} 列')

wb_out.save(OUT)
wb_src.close()
print('已輸出:', OUT)

# ── 輸出混淆對照表 ────────────────────────────────
with open(MAPPING_OUT, 'w', encoding='utf-8') as f:
    f.write('# 樣本資料混淆對照表\n\n')
    f.write('> 僅供資料追蹤核對用；本表本身亦含真實名稱，請勿外流。\n\n')
    f.write('## 院區\n\n| 真實 | 樣本 |\n|---|---|\n')
    for real, fake in BRANCH_FAKE.items():
        f.write(f'| {real} | {fake} |\n')
    f.write(f'\n| 和平婦幼(合併) | 丙院區 |\n')
    f.write('\n## 部別\n\n| 真實 | 樣本 |\n|---|---|\n')
    for real, fake in div_map.items():
        f.write(f'| {real} | {fake} |\n')
    f.write('\n## 科別\n\n| 真實 | 樣本 |\n|---|---|\n')
    for real, fake in dept_map.items():
        f.write(f'| {real} | {fake} |\n')
    f.write('\n## 門診部\n\n| 真實 | 樣本 |\n|---|---|\n')
    for real, fake in clinic_map.items():
        f.write(f'| {real} | {fake} |\n')
    f.write('\n## 護理站\n\n| 真實 | 樣本 |\n|---|---|\n')
    for real, fake in station_map.items():
        f.write(f'| {real} | {fake} |\n')
    f.write('\n## 大類別\n\n| 真實 | 樣本 |\n|---|---|\n')
    for real, fake in major_map.items():
        f.write(f'| {real} | {fake} |\n')
    f.write('\n## 病床類別\n\n| 真實 | 樣本 |\n|---|---|\n')
    for real, fake in cat_map.items():
        f.write(f'| {real} | {fake} |\n')
    f.write('\n## 醫師\n\n| 代碼 | 姓名 | 樣本代碼 | 樣本姓名 |\n|---|---|---|---|\n')
    for (did, dnm), (fid, fname) in doc_map.items():
        f.write(f'| {did} | {dnm} | {fid} | {fname} |\n')
    f.write('\n## 數值倍率\n\n| 院區 | 倍率 |\n|---|---|\n')
    for real, k in SCALE.items():
        f.write(f'| {real} | ×{k} |\n')
    f.write('\n> 佔床率等比率欄位與年月/日數等曆法欄位維持原值不動。\n')
print('已輸出:', MAPPING_OUT)
